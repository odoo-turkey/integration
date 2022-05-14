# Copyright 2022 Yiğit Budak (https://github.com/yibudak)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

from odoo import models, fields, api
import openpyxl
from pathlib import Path

lower_map = {
    ord(u'I'): u'ı',
    ord(u'İ'): u'i',
    }


class AddressDistrict(models.Model):
    _inherit = "address.district"

    sendeo_code = fields.Char(string='Sendeo District Code')

    @api.model
    def import_sendeo_district_codes(self):
        # TODO: This method is too risky. It should be reviewed or removed.
        module_path = Path(__file__).parent.parent.absolute()
        file_path = 'data/il_ilce_kodlari.xlsx'
        file = openpyxl.load_workbook(Path(module_path, file_path))
        sheet = file.active
        if not sheet.title == 'İlçe Kodları':
            return False

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            state_code = "0%s" % row[1].value if len(str(row[1].value)) == 1 else row[1].value
            district_name = row[2].value.translate(lower_map)
            if 'MERKEZ' in district_name:
                district_name = 'Merkez'

            if '/' in district_name:
                district_name = district_name.split('/')[0]

            district = self.search([('name', 'ilike', district_name),
                                    ('state_id.code', '=', state_code),
                                    ('sendeo_code', '=', False)])
            if district:
                fields.first(district).sendeo_code = row[0].value
